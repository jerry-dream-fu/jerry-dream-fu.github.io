# -*- coding: utf-8 -*-
from email.mime.application import MIMEApplication
from itertools import count
import urllib
from urllib.request import urlopen
from bs4 import BeautifulSoup
from bs4.element import Tag 
import openpyxl
from openpyxl import Workbook
import datetime
import time
from openpyxl.comments import Comment
import os
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
from typing import ContextManager
import requests
#from win32com.client import Dispatch

today = datetime.datetime.now().strftime("%Y-%m-%d")



current_path = os.getcwd()
save_file = current_path+'/teacher_info_30000_40000.xlsx'

#def just_open(filename):
#    xlApp = Dispatch("Excel.Application")
 #   xlApp.Visible = False
#    xlBook = xlApp.Workbooks.Open(filename)
 #   xlBook.Save()
 #   xlBook.Close()




xlsx_list=['姓名','性别','所属院校','所属院系','职称','导师类型','招生专业','通讯方式','个人简述','科研工作','教育背景']

wb = 1
worksheet_1 = 1
worksheet_2_chinakaoyan = 1
if os.path.exists(save_file):#true 
    #global worksheet_1
    #global worksheet_2_chinakaoyan
    #global wb
    print("excel exists")
    wb = openpyxl.load_workbook(save_file,data_only=True)
    sheet_names = wb.sheetnames
    worksheet_1 = wb[sheet_names[0]]#
else:
    wb = Workbook()
    worksheet_1 = wb.active
    worksheet_1.title= '导师信息'
    worksheet_1.append(xlsx_list)

    
#three sheet


    
    

def get_teacher_info_xmc(url):  
    headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36'}
    req=urllib.request.Request(url,headers=headers)
    res=urlopen(req)
    bsObj = BeautifulSoup(res,features="lxml")
    #body > div.bg > div:nth-child(9) > div:nth-child(2) > table > tbody.forum_body_manage
    #body > div.bg > div:nth-child(9) > div:nth-child(2) > table > tbody.forum_body_manage > tr:nth-child(8) > td.xmc_lp20

    text_list=[]
    tea_basic_info = bsObj.find_all("div",{"class":"teacher-td"})## 一页这一个tag
    tea_detail_info = bsObj.find_all("div",{"class":"lf-item"})## 一页这一个tag
    if len(tea_basic_info) == 0:
        return
    #print(tea_basic_info)
    #print(type(tea_basic_info))#<class 'bs4.element.ResultSet'>
    teacher_name_and_sex = tea_basic_info[0]#导师姓名
    teacher_name = teacher_name_and_sex.find_all("div")[0].get_text().split("：")[1]
    teacher_sex = teacher_name_and_sex.find_all("div")[1].get_text().split("：")[1]

    teacher_uni_and_school = tea_basic_info[1]#所属学校
    teacher_uni = teacher_uni_and_school.find_all("div")[0].get_text().strip().split("：")[1]
    teacher_school = teacher_uni_and_school.find_all("div")[0].next_sibling.get_text().strip().split("：")[1]
    print(len(tea_basic_info[1]))

    teacher_title_and_stu = tea_basic_info[2]#职称
    teacher_title = tea_basic_info[2].find_all("div")[0].get_text().strip().split("：")[1]#
    teacher_stu = tea_basic_info[2].find_all("div")[0].next_sibling.get_text().strip().split("：")[1]#

    teacher_direction = tea_basic_info[3].get_text().strip().split("：")[1]##招生专业

    print(teacher_name)
    print(teacher_sex)
    print(teacher_uni)
    print(teacher_school)
    print(teacher_title)
    print(teacher_stu)
    print(teacher_direction)
    text_list.append(teacher_name)
    text_list.append(teacher_sex)
    text_list.append(teacher_uni)
    text_list.append(teacher_school)
    text_list.append(teacher_title)
    text_list.append(teacher_stu)
    text_list.append(teacher_direction)

    phone_and_email_info=''
    personal_intro=''
    research_work=''
    tea_edu_background=''
    for index in range(len(tea_detail_info)):
        text = tea_detail_info[index].get_text().strip().replace('\n',' ').replace('\t','')
        #print(text)
        if phone_and_email_info is '':
            if '通讯方式' in text:
                phone_and_email_info = text.split(":")[1].strip()
                print("phone_and_email_info" +phone_and_email_info)

        if personal_intro is '':
            if '个人简述' in text:
                personal_intro = text.split(":")[1].strip()

        if research_work is '':
            if '科研工作' in text:
                research_work = text.split(":")[1].strip()
        if research_work is '':
            if '教育背景' in text:
                tea_edu_background = text.split(":")[1].strip()


    print("phone_and_email_info" +phone_and_email_info)
    text_list.append(phone_and_email_info)   
    text_list.append(personal_intro) 
    text_list.append(research_work) 
    text_list.append(tea_edu_background) 
    #if 0 < len(tea_detail_info):#通讯方式
        #phone_and_email_info = tea_detail_info[0].get_text().strip().replace('\n',' ').replace('\t','')
    #if 1 < len(tea_detail_info):#个人简述
    #    personal_intro = tea_detail_info[1].get_text().strip().replace('\n',' ').replace('\t','')
        #print(phone_and_email_info)
    #if 2 < len(tea_detail_info):#科研工作
        #research_work = tea_detail_info[2].get_text().strip().replace('\n',' ').replace('\t','')       
        #tea_edu_background = tea_detail_info[2].get_text().strip().replace('\n',' ').replace('\t','')
    #教育背景
        #print(tea_edu_background)
    worksheet_1.append(text_list)
    wb.save(filename=save_file)



    




if __name__=="__main__":
    #
    url = "https://daoshi.eol.cn/tutor/{}"
    urls = [url.format(str(i)) for i in range(30000,40000)]#
    count=30000
    for url in urls:
        count=count+1
        get_teacher_info_xmc(url)
        if(count%1000==0):
            time.sleep(5)
        print(count)
    

    

