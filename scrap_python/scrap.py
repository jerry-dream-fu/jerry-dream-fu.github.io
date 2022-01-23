from email.mime.application import MIMEApplication
from itertools import count
from tokenize import cookie_re
import urllib
from urllib.request import urlopen
from bs4 import BeautifulSoup
from bs4.element import Tag 
import openpyxl
from openpyxl import Workbook
import datetime
import time
from openpyxl.comments import Comment
# -*- coding: utf-8 -*-
import os
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
from typing import ContextManager
import requests

##
##小木虫
##    https://mapi.xmcimg.com/bbs/kaoyan.php?emobile=3
##中国考研网
##    http://www.chinakaoyan.com/tiaoji/schoollist.shtml
##中国教育在线     http://www.eol.cn/html/ky/kytj/info.shtml
# save to excel
#get today info
today = datetime.datetime.now().strftime("%Y-%m-%d")



current_path = os.getcwd()
save_file = current_path+'/info.xlsx'

#wb = 1
worksheet_1 = 1
#worksheet_2_chinakaoyan = 1

xlsx_list=['标题','学校','专业','调剂人数','发布时间','链接']
#if os.path.exists(save_file):#true 
    #global worksheet_1
    #global worksheet_2_chinakaoyan
    #global wb
#    wb = openpyxl.load_workbook(save_file)
#    sheet_names = wb.sheetnames
#    worksheet_1 = wb[sheet_names[0]]#'小木虫'
#    worksheet_2_chinakaoyan=wb[sheet_names[1]]#'中国考研网'
#else:
    #global worksheet_1
    #global worksheet_2_chinakaoyan
    #global wb
wb = Workbook()
worksheet_1 = wb.active
worksheet_1.title= '小木虫'
worksheet_1.append(xlsx_list)
worksheet_2_chinakaoyan=wb.create_sheet('中国考研网')
worksheet_2_chinakaoyan.append(xlsx_list)
    
#three sheet



xmu_count=0
def get_info_xmc(url):  
    headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36'}
    req=urllib.request.Request(url,headers=headers)
    res=urlopen(req)
    bsObj = BeautifulSoup(res,features="lxml")
    #body > div.bg > div:nth-child(9) > div:nth-child(2) > table > tbody.forum_body_manage
    #body > div.bg > div:nth-child(9) > div:nth-child(2) > table > tbody.forum_body_manage > tr:nth-child(8) > td.xmc_lp20
    info_list = bsObj.findAll("tbody",{"class":"forum_body_manage"})## 一页这一个tag
    #print(len(info_list))
    #info in full page
    for l in info_list:
        #a item in info
        for l_child in l.children:
            if isinstance(l_child, Tag):
                text_list=[]
                #print(type(l_child))#<class 'bs4.element.Tag'>
                #a item
                l_child_child = l_child.find_all('td')
                #print(len(l_child_child))
                tilte_and_link = l_child_child[0]
                title_info = tilte_and_link.get_text()
                text_list.append(title_info)#'标题'
                #print()
                link_info = tilte_and_link.find('a').get('href')
                #print(type(link_info))
                shool_info = l_child_child[1].get_text()
                major_info = l_child_child[2].get_text()
                num_info = l_child_child[3].get_text()
                release_time = l_child_child[4].get_text()

                get_y_m_d = release_time.split()[0]
                if get_y_m_d == today:
                    xmu_count=xmu_count+1
                if True:#
                    global xmu_count
                    xmu_count=xmu_count+1
                    text_list.append(shool_info)#,'学校'
                    text_list.append(major_info)#'专业'
                    text_list.append(num_info)#'调剂人数
                    text_list.append(release_time)#'发布时间'
                    text_list.append(link_info)#'链接'
                    worksheet_1.append(text_list)
    
    wb.save(filename=save_file)


chinakaoyan_count=0
def get_info_chinakaoyan(url):
    headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36'}
    req=urllib.request.Request(url,headers=headers)
    res=urlopen(req)
    bsObj = BeautifulSoup(res,features="lxml")
    #body > div.bg > div:nth-child(9) > div:nth-child(2) > table > tbody.forum_body_manage
    #body > div.bg > div:nth-child(9) > div:nth-child(2) > table > tbody.forum_body_manage > tr:nth-child(8) > td.xmc_lp20
    info_list = bsObj.findAll("div",{"class":"info-item font14"})## 一页这一个tag
    #print(type(info_list))
    #info in full page
    for info in info_list:
        #print(type(info))
        text_list=[]
        school_info = info.find('span',class_='school')
        name_info = info.find('span',class_='name')
        title_link_info = info.find('span',class_='title')
        link_info=''
        if link_info != None:
            link_info = title_link_info.find('a').get('href')
        release_time_info = info.find('span',class_='time')
        get_y_m_d = release_time_info.get_text().split()[0]
       # print(get_y_m_d)
       if get_y_m_d == today:
           chinakaoyan_count=chinakaoyan_count+1
        if True:#
            global chinakaoyan_count
            chinakaoyan_count=chinakaoyan_count+1        
            text_list.append(title_link_info.get_text())#'标题',
            text_list.append(school_info.get_text())#'学校',
            text_list.append(name_info.get_text())#'专业',
            text_list.append('')#'调剂人数',
            text_list.append(release_time_info.get_text())#'发布时间',
            text_list.append('http://www.chinakaoyan.com/'+link_info)#'链接'
            worksheet_2_chinakaoyan.append(text_list)
  
    wb.save(filename=save_file)
                
#worksheet_eol=wb.create_sheet('中国教育在线')
#worksheet_eol.append(xlsx_list)
def get_info_eol(url):
    headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36'}
    req=urllib.request.Request(url,headers=headers)
    res=urlopen(req)
    bsObj = BeautifulSoup(res,features="lxml")
    #body > div.bg > div:nth-child(9) > div:nth-child(2) > table > tbody.forum_body_manage
    #body > div.bg > div:nth-child(9) > div:nth-child(2) > table > tbody.forum_body_manage > tr:nth-child(8) > td.xmc_lp20
    info_list = bsObj.findAll("div",{"class":"info-item font14"})## 一页这一个tag
    #print(type(info_list))
    #info in full page
    for info in info_list:
        #print(type(info))
        text_list=[]
        school_info = info.find('span',class_='school')
        name_info = info.find('span',class_='name')
        title_link_info = info.find('span',class_='title')
        link_info=''
        if title_link_info != None:
            link_info = title_link_info.find('a').get('href')
        release_time_info = info.find('span',class_='time')
        
        text_list.append(title_link_info.get_text())#'标题',
        text_list.append(school_info.get_text())#'学校',
        text_list.append(name_info.get_text())#'专业',
        text_list.append('')#'调剂人数',
        text_list.append(release_time_info.get_text())#'发布时间',
        text_list.append('http://www.chinakaoyan.com/'+link_info)#'链接'
        worksheet_chinakaoyan.append(text_list)


    wb.save(filename=save_file)

mailHost = 'smtp.163.com'
mailPort = 465

user_lxf='lxf1632046131@163.com'
passw_lxf = 'ZVENGRMQAKMXXYUC'
#os.environ.get("ZVENGRMQAKMXXYUC")
# passw_lxf = os.environ.get("lixiangfu@146")

def sendMail(content):
    receiver=['755438454@qq.com','1632046131@qq.com','jzsmail@163.com']#
    smptp = smtplib.SMTP_SSL(mailHost,mailPort)
    smptp.login(user=user_lxf,password=passw_lxf)

    msg = MIMEMultipart()

    msg['Subject']=Header("调剂信息",'utf-8')
    msg['from']=user_lxf
    msg['to']=','.join(receiver)
    
    #print(content)
    msg_content=MIMEText(content,'plain','utf-8')
    msg.attach(msg_content)
    
    print("准备添加附件...")
    part = MIMEApplication(open(current_path+'/info.xlsx','rb').read())
    part.add_header('Content-Disposition', 'attachment', filename="info.xlsx")#给附件重命名,一般和原文件名一样,改错了可能无法打开.
    msg.attach(part)

    smptp.sendmail(user_lxf,receiver,msg.as_string())

if __name__=="__main__":
    #
    url = "http://muchong.com/bbs/kaoyan.php?&page={}"
    urls = [url.format(str(i)) for i in range(1,500)]
    for url in urls:
        get_info_xmc(url)
    xmu_content = "小木虫 "+ today+" 更新调剂条目条数： " + str(xmu_count)
    print(xmu_content)

    url_chinakaoyan = "http://www.chinakaoyan.com/tiaoji/schoollist/pagenum/{}.shtml"
    urls_chinakaoyan = [url_chinakaoyan.format(str(i)) for i in range(1,100)]
    for url in urls_chinakaoyan:
        get_info_chinakaoyan(url)
    chain_kaoyan_content = "中国考研网 "+ today+" 更新调剂条目条数： " + str(chinakaoyan_count)
    print(chain_kaoyan_content)

    #print(type(today))
    content = today +" 更新调剂条目条数： " + str(xmu_count + chinakaoyan_count)
    print(content)
    if xmu_count + chinakaoyan_count > 0:
        sendMail(content+': '+xmu_content+' '+chain_kaoyan_content)
    

